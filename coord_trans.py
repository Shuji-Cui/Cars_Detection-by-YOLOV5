import gdal
import os
import pathlib
import glob
import re

import openpyxl


def increment_path(name):
    """
    让保存文件的序号逐步增加
    如：".txt1"-->".txt2"-->".txt3"
    """
    root = "data/coords_output/output"
    path = pathlib.Path(root) / name  # 转换为可操作的系统路径
    if path.exists():
        suffix = path.suffix   # 获取后缀名
        path = path.with_suffix('')
        dirs = glob.glob(f"{path}*")  # *代表获取所有符合的文件
        matches = [re.search(rf"%s(\d+)" % path.stem, d) for d in dirs]   # 在d中搜索符合正则表达式的对象
        i = [int(m.groups()[0]) for m in matches if m]  # indices
        n = max(i) + 1 if i else 2  # increment number    获取的序号是当前最大值，下一个序号是当前最大值+1
        path = pathlib.Path(f"{path}{n}{suffix}")  # update path
    dir = path if path.suffix == '' else path.parent  # directory
    if not dir.exists():
        dir.mkdir(parents=True, exist_ok=True)  # make directory
    return path

def coord_transor(idx, coord_pixel, img_path):
    """
    coord_pixel:像素坐标对
    save_path:坐标保存的位置
    imgs_path:图片路径

    关于仿射矩阵信息：
    0：图像左上角的X坐标；
    1：图像东西方向分辨率；
    2：旋转角度，如果图像北方朝上，该值为0；
    3：图像左上角的Y坐标；
    4：旋转角度，如果图像北方朝上，该值为0；
    5：图像南北方向分辨率；
    """

    # 打开图片
    img = gdal.Open(img_path)
    # 获取当前图片为第几行第几列
    # row_col = idx.split(".")[0]
    row = idx.split("_")[0]  # 行
    col = idx.split("_")[1]  # 列
    # 获取仿射矩阵信息
    geomessage = img.GetGeoTransform()
    # 生成对应的地理坐标
    xgeo = geomessage[0]+geomessage[1]*(coord_pixel[0]+(int(col)-1)*768) + ((int(row)-1)*768+coord_pixel[1])*geomessage[2]
    ygeo = geomessage[3]+geomessage[5]*(coord_pixel[1]+(int(row)-1)*768) + ((int(col)-1)*768+coord_pixel[0])*geomessage[4]

    return xgeo, ygeo

def creat_wb():
    """
    创建一个xlsx接收坐标
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    ws["A1"] = "xgeo"
    ws["B1"] = "ygeo"

    return wb

def coord_output(wb, xgeo, ygeo, i):
    """
    将坐标输出到xlsx
    wb:输出的xlsx
    xgeo:x地理坐标
    ygeo:y地理坐标
    i:行号

    """
    ws = wb.active
    ws[f"A{i}"] = xgeo
    ws[f"B{i}"] = ygeo




if "__main__":
    pass